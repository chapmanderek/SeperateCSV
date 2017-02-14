# CVS Seperater

# Todo
# change user input so it only repeats the section that has a problem
# allow importing of excel files - currently doesnt handle fields with a comma
# ask for user input on output settings, # of pages wide etc
# add date stamp to header or footer

import datetime as dt
import os
import xlwt
import openpyxl as opx
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side

header_style = NamedStyle(name='header_style')
header_style.font = Font(bold=True, size=12, name='Arial')
header_style.alignment = Alignment(horizontal='center')
header_style.fill = PatternFill('solid', 'DCDCDC')
header_style.border = Border(bottom=Side(border_style='thin'))

row_style = NamedStyle(name='row_style')
row_style.font = Font(bold=True, size=12, name='Trebuchet')
row_style.alignment = Alignment(horizontal='center')
row_style.alignment = Alignment(horizontal='left')
row_style.border = Border(bottom=Side(border_style='thin'))

def open_py_output(filename, header, data):
	wbook = opx.Workbook()
	wsheet = wbook.active
	wbook.add_named_style(header_style)
	wbook.add_named_style(row_style)

	header_parts = [ each.lstrip().rstrip() for each in header.split(',') ]
	header_width = len(header_parts)
	wsheet.append(header_parts)

	for row in wsheet.iter_rows():
		for cell in row:
			cell.style = header_style

	# split data list into individual rows then append each on to the spreadsheet
	row_count = 1
	for row in data:
		row_parts = [ each.lstrip().rstrip() for each in row.split(',') ]
		wsheet.append(row_parts)
		row_count += 1

	# add stle for data rows
	for row in wsheet.iter_rows(min_row=2, max_row=row_count):
		for cell in row:
			cell.style = row_style
	
	for column in wsheet.columns:
		col_max_len = max( [len(cell.value) for cell in column if cell.value is not None] )

		if col_max_len < 10 : col_max_len = 10
		if col_max_len > 30 : col_max_len = 30
		wsheet.column_dimensions[column[0].column].width = col_max_len*1.1

	# put date stamp in right header   &D
	wsheet.oddHeader.right.text = "&D"
	wsheet.oddHeader.right.font_size = 14

	# worksheet print options
	wsheet.print_options.horizontalcentered = True
	wsheet.page_setup.orientation = "landscape"
	wsheet.page_setup.fitToPage = True

	wbook.save(filename)
	return (row_count-1)

# get filename
print("Thanks for using this file splitter.  Use q to quit.")
keep_going = False
while(keep_going == False):
	keep_going = True
	file = input('Enter csv name:')
	if file == '' : file = 'test_data.csv'
	if file == 'q' or file == 'Q' : exit()
	
	try:
		csv_handle = open(file)
		# read in the header but get rid of trailing new line characters
		header = csv_handle.readline().rstrip()
	except:
		keep_going = False
		print('\nSomethings wrong with that filename.  Try again or use q to quit.')
		continue


	# get column to seperate the file by from user error check so exits gracefully
	seperate_by = input('Enter column to seperate by (default is 1):')
	if file == 'q' or file == 'Q' : exit()
	if seperate_by == '' : seperate_by = 1
	elif seperate_by.isdigit() == True : seperate_by = int(seperate_by)

	else:
		print('Not valid input for some reason')
		keep_going = False
		continue

	# assume user means 1 index values for the column number
	# make sure the number is a valid column and exit gracefully if not
	seperate_by -= 1
	print('{0} --> user input   {1} --> num columns'.format(seperate_by, len(header.split(','))))
	if len(header.split(',')) < seperate_by or seperate_by < 0:
		print('\nThe number you entered to seperate the worksheet by was no good. Try again or use q to quit.')
		keep_going = False
		continue


# seperate the rest of the info into a dict of lists of strings
sorted_info = dict()
lines_count= 0
for line in csv_handle:
	lines_count += 1
	line = line.rstrip()
	line_parts = line.split(',')
	line_parts = [each.lstrip() for each in line_parts]
	
	if line_parts[seperate_by] in sorted_info.keys():
		sorted_info[line_parts[seperate_by]].append(line)
	else:
		sorted_info[line_parts[seperate_by]] = [line]

# setup a date stamp for the folder and individual worksheets / create folder
today = dt.date.today()
formatted_date = "{month}-{day}-{year}".format(month=today.month, day=today.day, year=today.year)
dir_name = '{cwd}/{date}'.format(date=formatted_date, cwd=os.getcwd())
if not os.path.exists(dir_name) : os.mkdir(dir_name)

# create the individual sheets
rows_written = 0
sheets_created = 0
for key in sorted_info.keys():
	keyname = key
	if keyname == '' : keyname = 'Blanks'
	file_name = "{dn}/{key}_{date}.xlsx".format(key=keyname, date=formatted_date, dn=dir_name)
	rows_written += open_py_output(file_name, header, sorted_info[key])
	sheets_created += 1

print('\n{0} records were read in.'.format(lines_count))
print('{0} records were written to {1} different sheets. \n'.format(rows_written, sheets_created))

# if there were any blanks in the records print how many were found
num_blanks = 0
if '' in sorted_info.keys() : num_blanks = len(sorted_info[''])
if num_blanks != 0:
	print('NOTE: {0} records with blank key columns were found and written to a file named Blanks.'.format(num_blanks))


